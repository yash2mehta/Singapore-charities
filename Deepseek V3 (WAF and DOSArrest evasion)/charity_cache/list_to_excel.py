#!/usr/bin/env python3
"""
list_to_excel.py
=================
Builds a clean, formatted .xlsx from a cached charities.gov.sg list JSON file
(e.g. list_registered.json, list_ipc.json, ...) -- the kind of file the
charity_scraper.py `scrape`/`all` commands save under ./charity_cache/.

This works from the LIST-level cache only (no Organisation Profile / Financial
Information detail panels needed) -- so it runs instantly, offline, on
whatever has been scraped so far, even a partial/resumed run.

Usage
-----
    python list_to_excel.py list_registered.json
    python list_to_excel.py list_registered.json --out Registered_Charities.xlsx

Output
------
One sheet, "Charities", one row per charity:
  - Name, UEN, Status, Registration/De-registration dates, Charity Setup
  - Sector, Sector Administrator, Classification(s), Activities
  - IPC status/validity
  - Website fields (NameofURL / OtherURL) placed near the front, not buried
  - Any field present in the JSON that isn't explicitly mapped above still
    shows up as an Extra_* column -- nothing is silently dropped.
Columns auto-width, header row frozen/bold, list-valued fields (Activities,
PrimaryClassification) rendered as clean multi-line text, not JSON.
"""

from __future__ import annotations

import argparse
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Fields we handle explicitly, in the order they should appear. Each maps to
# (header label, how to pull/format the value from a record dict).
FIELD_ORDER = [
    ("CharityIPCName", "Charity Name"),
    ("UENNo", "UEN No"),
    ("CharityStatus", "Status"),
    ("CharitySetup", "Charity Setup"),
    ("RegistrationDate", "Registration Date"),
    ("DeRegistrationDate", "De-Registration Date"),
    ("DeRegistrationText", "De-Registration Note"),
    ("NameofURL", "Name of URL"),
    ("OtherURL", "Other URL"),
    ("PrimarySector", "Primary Sector"),
    ("SectorAdministrato", "Sector Administrator"),
    ("PrimaryClassification", "Classification(s)"),
    ("Activities", "Activities"),
    ("IPCStatus", "IPC Status"),
    ("IPCValidFrom", "IPC Valid From"),
    ("IPCValidTill", "IPC Valid Till"),
    ("shoIPCPeriod", "Show IPC Period"),
    ("Type", "Type"),
    ("CharityAccountCRMRecordID", "Charity Record GUID"),
]
HANDLED_KEYS = {k for k, _ in FIELD_ORDER} | {"TotalRecords", "RelatedAccountOrgID"}

MULTILINE_HEADERS = {"Classification(s)", "Activities"}


def clean_text(v):
    """Strings from this API often carry stray leading/trailing whitespace or
    newlines (e.g. CharityIPCName). Also renders None/list values sensibly."""
    if v is None:
        return ""
    if isinstance(v, list):
        return "\n".join(clean_text(item) for item in v if item not in (None, ""))
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return str(v).strip()


def load_records(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        # tolerate being pointed at a raw API response instead of the cache list
        data = data.get("charityInfosData") or data.get("data") or []
    if not isinstance(data, list):
        raise ValueError(f"Expected a JSON list of charity records in {path}, got {type(data)}")
    return data


def collect_extra_columns(records):
    """Any key present in the data that isn't in HANDLED_KEYS gets its own
    Extra_<field> column, so nothing from the source JSON is silently dropped
    even if the API adds fields later."""
    extras = []
    seen = set()
    for rec in records:
        for k in rec.keys():
            if k not in HANDLED_KEYS and k not in seen:
                seen.add(k)
                extras.append(k)
    return extras


def build_workbook(records, source_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Charities"

    extra_keys = collect_extra_columns(records)
    columns = [label for _, label in FIELD_ORDER] + [f"Extra_{k}" for k in extra_keys]

    header_font = Font(bold=True)
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
    ws.freeze_panes = "A2"

    max_len = [len(h) for h in columns]
    max_lines = [1] * len(columns)

    for row_idx, rec in enumerate(records, start=2):
        for col_idx, (key, label) in enumerate(FIELD_ORDER, start=1):
            value = clean_text(rec.get(key))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if label in MULTILINE_HEADERS and "\n" in value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                max_lines[col_idx - 1] = max(max_lines[col_idx - 1], value.count("\n") + 1)
            longest_line = max((len(line) for line in value.split("\n")), default=0)
            max_len[col_idx - 1] = max(max_len[col_idx - 1], longest_line)

        for offset, key in enumerate(extra_keys):
            col_idx = len(FIELD_ORDER) + offset + 1
            value = clean_text(rec.get(key))
            ws.cell(row=row_idx, column=col_idx, value=value)
            max_len[col_idx - 1] = max(max_len[col_idx - 1], len(value))

        # Row height: fits the tallest multi-line field in this row (~15pt/line),
        # so wrapped Activities/Classification lists aren't clipped.
        lines_in_row = max(
            (clean_text(rec.get(key)).count("\n") + 1 for key in MULTILINE_HEADERS_KEYS(FIELD_ORDER, MULTILINE_HEADERS)),
            default=1,
        )
        if lines_in_row > 1:
            ws.row_dimensions[row_idx].height = min(15 * lines_in_row, 300)

    for col_idx, width in enumerate(max_len, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)

    ws.auto_filter.ref = ws.dimensions

    # A small "Source" note sheet -- where this came from and how many rows,
    # useful since this is built from a possibly-partial resumable cache.
    note = wb.create_sheet("Source")
    note["A1"] = "Source cache file"
    note["B1"] = str(source_path)
    note["A2"] = "Records exported"
    note["B2"] = len(records)
    note["A1"].font = note["A2"].font = Font(bold=True)

    return wb


def MULTILINE_HEADERS_KEYS(field_order, multiline_headers):
    return [key for key, label in field_order if label in multiline_headers]


def main(argv=None):
    parser = argparse.ArgumentParser(description="Build an .xlsx from a cached charities.gov.sg list JSON file.")
    parser.add_argument("json_path", help="Path to the cached list JSON (e.g. list_registered.json).")
    parser.add_argument("--out", default=None, help="Output .xlsx path (default: <input name>.xlsx).")
    args = parser.parse_args(argv)

    records = load_records(args.json_path)
    if not records:
        print(f"No records found in {args.json_path}.")
        return 1

    out_path = args.out or (args.json_path.rsplit(".", 1)[0] + ".xlsx")
    wb = build_workbook(records, args.json_path)
    wb.save(out_path)
    print(f"Wrote {len(records)} charities -> {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
