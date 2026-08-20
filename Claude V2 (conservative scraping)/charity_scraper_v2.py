#!/usr/bin/env python3
"""
charities.gov.sg Advance Search scraper
=======================================

Scrapes the Singapore Charity Portal (https://www.charities.gov.sg/Pages/AdvanceSearch.aspx)
for each "Status of Charity" and enriches every charity with its
**Organisation Profile** and **Financial Information** detail pages, then writes
the combined result to Excel.

How it works (confirmed from the site's network traffic)
--------------------------------------------------------
1. The results list comes from a POST to:
       /_layouts/15/CPInternet/AdvanceSearchHandler.ashx
   It's a jQuery DataTables request. The status filter lives inside the
   `query` field as `?advType=0&type=<CODE>` and pagination is `start`/`length`.
   Each returned record carries `CharityAccountCRMRecordID` -- a GUID.

2. When you click a charity, its detail panels come from:
       /_layouts/15/CPInternet/SearchResultHandler.ashx?query=<Q>&type=<PANEL>
   where <Q> is simply base64(GUID) and <PANEL> is
   "Organisation Profile" or "Financial Information".

So the whole job is: list -> collect GUIDs -> base64 -> pull the two panels.

Pipeline (sub-commands)
-----------------------
    selftest   Offline check that URL/base64 logic matches a known example. No network.
    discover   Probe candidate `type` codes and print which CharityStatus each returns,
               so you can fill STATUS_TYPE_CODES with confidence.
    scrape     Fetch the list for a status (with GUIDs) + both detail panels for every
               charity, caching every raw JSON response to disk. Resumable.
    export     Read the cache and write an enriched .xlsx. Re-runnable without re-scraping.
    all        scrape + export for one status.

Typical use
-----------
    python charity_scraper.py selftest
    python charity_scraper.py discover
    python charity_scraper.py all --status registered
    python charity_scraper.py all --status ipc
    python charity_scraper.py all --status deregistered
    python charity_scraper.py all --status exempt
    python charity_scraper.py all --status deexempted

Notes
-----
* No login / cookies are required. The site sits behind DOSarrest (a CDN/DDoS
  layer), so the script uses a realistic browser header set, a polite delay
  between requests, and exponential-backoff retries. Keep --delay >= 0.4 to be
  a good citizen and avoid being throttled.
* The scrape is fully resumable: every response is cached under ./charity_cache/.
  Re-running skips whatever is already on disk. Delete a file (or use --refresh)
  to re-fetch it.
* The exact JSON shape of the two detail panels was not in the captured traffic,
  so `export` flattens whatever structure comes back and also keeps the raw JSON
  in a column. Once you know the real shape, tailoring the columns is trivial --
  the raw cache means you never have to re-scrape to change the Excel layout.
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import random
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import quote

import requests

# Force every print() in this script to flush immediately. Without this, output
# can sit in a buffer and make a perfectly-running (but slow / retrying) scrape
# look frozen -- especially on Windows consoles or if output is piped/logged.
import builtins as _builtins
def print(*args, **kwargs):  # noqa: A001 - intentional shadow, module-local only
    kwargs.setdefault("flush", True)
    _builtins.print(*args, **kwargs)

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

BASE = "https://www.charities.gov.sg"
SEARCH_URL = BASE + "/_layouts/15/CPInternet/AdvanceSearchHandler.ashx"
DETAIL_URL = BASE + "/_layouts/15/CPInternet/SearchResultHandler.ashx"

CACHE_DIR = "charity_cache"

# The `type` code inside the search `query` string selects the Status of Charity.
# Registered = 100000000 is CONFIRMED from the captured request. The others are
# best-guess sequential values -- VERIFY them once with the `discover` command
# (it prints the CharityStatus each code actually returns) and correct as needed.
STATUS_TYPE_CODES = {
    "registered":   100000000,   # CONFIRMED
    "ipc":          100000001,   # verify with `discover`
    "deregistered": 100000002,   # verify with `discover`
    "exempt":       100000003,   # verify with `discover`
    "deexempted":   100000004,   # verify with `discover`
}

# Friendly labels used in filenames / messages.
STATUS_LABELS = {
    "registered":   "Registered Charities",
    "ipc":          "IPCs",
    "deregistered": "De-registered Charities",
    "exempt":       "Exempt Charities",
    "deexempted":   "De-exempted Charities",
}

# The two detail panels to pull for each charity.
DETAIL_PANELS = {
    "profile":   "Organisation Profile",
    "financial": "Financial Information",
}

# Browser-like headers (mirrors the captured request).
HEADERS = {
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "en-US,en;q=0.9",
    "Origin": BASE,
    "Referer": BASE + "/Pages/AdvanceSearch.aspx",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "X-Requested-With": "XMLHttpRequest",
}

# --------------------------------------------------------------------------- #
# HTTP helpers
# --------------------------------------------------------------------------- #

def make_session(pool_size: int = 16) -> requests.Session:
    """requests.Session is safe to share across threads for concurrent requests
    (we only read its headers, never mutate them mid-flight). Size the connection
    pool to the worker count so urllib3 doesn't discard/recreate connections.

    Using ONE session for the whole run matters: it keeps the same TCP/TLS
    connections and the same cookie jar alive, which is both cheaper for the
    server and how a normal client behaves."""
    s = requests.Session()
    s.headers.update(HEADERS)
    adapter = requests.adapters.HTTPAdapter(pool_connections=pool_size, pool_maxsize=pool_size)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s


def warm_up_session(session, timeout=30, quiet=False):
    """Load the actual Advance Search page once so the session holds whatever
    cookies the site hands out before we start calling its XHR handlers.

    This is ordinary session hygiene, not a trick: calling an AJAX endpoint with
    no prior page load and an empty cookie jar is unusual traffic, and reloading
    the page after a block is also how a real user recovers. Non-fatal on failure.
    """
    try:
        session.get(
            BASE + "/Pages/AdvanceSearch.aspx",
            timeout=timeout,
            headers={
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "X-Requested-With": None,   # None removes the header for this request
                "Origin": None,
            },
        )
        if not quiet:
            print(f"  session: warmed up on the search page ({len(session.cookies)} cookie(s) held).")
    except Exception as exc:
        if not quiet:
            print(f"  session: warm-up request failed ({exc}) -- continuing anyway.")


class RateLimiter:
    """ONE global request budget shared by every thread.

    This is the single most important change versus the previous version. Pacing
    used to be per-thread (--delay slept inside each worker), so the rate actually
    hitting the server was workers/delay: 3 workers at --delay 1.0 is ~3 req/s,
    roughly an order of magnitude more than this site tolerates. Here every request
    in the process -- list pages and detail panels alike -- takes a numbered slot
    from one shared schedule, so the aggregate rate is what you asked for no matter
    how many workers are running.

    It also adapts: sustained pushback stretches the interval, and a long clean run
    slowly relaxes it back toward the target.
    """

    def __init__(self, per_minute: float = 24.0, floor_seconds: float = 0.0):
        base = 60.0 / per_minute if per_minute > 0 else 0.0
        self.base_interval = max(base, floor_seconds)
        self.interval = self.base_interval
        self.max_interval = max(self.base_interval * 12, 30.0)
        self._lock = threading.Lock()
        self._next_slot = 0.0
        self._clean_streak = 0

    def acquire(self):
        """Block until this thread's turn. Slots are handed out under the lock and
        slept on outside it, so N threads queue instead of colliding."""
        with self._lock:
            now = time.monotonic()
            # Jitter every gap: a fixed exact interval is both unlike real traffic
            # and prone to resonating badly with a fixed-window rate limiter.
            gap = self.interval * random.uniform(0.75, 1.45)
            slot = max(now, self._next_slot)
            self._next_slot = slot + gap
        wait = slot - time.monotonic()
        if wait > 0:
            time.sleep(wait)

    def pause(self, seconds: float):
        """Push the shared schedule forward so ALL threads wait, not just the one
        whose request was rejected."""
        with self._lock:
            self._next_slot = max(self._next_slot, time.monotonic() + seconds)

    def report_throttled(self, factor: float = 1.6):
        """Server pushed back -> slow the steady-state rate (up to a cap) for the
        rest of the run. Running slower for an hour beats getting blocked every
        250 rows and re-running five times."""
        with self._lock:
            self._clean_streak = 0
            new_interval = min(self.interval * factor, self.max_interval)
            changed = (new_interval - self.interval) > 0.05
            self.interval = new_interval
            shown = self.interval
        if changed:
            print(f"  rate: easing down to ~{60.0 / shown:.1f} requests/min "
                  f"({shown:.1f}s apart) after server pushback.")

    def report_success(self):
        """After a long clean stretch, creep back toward the requested rate."""
        recovered = None
        with self._lock:
            self._clean_streak += 1
            if self._clean_streak >= 60 and self.interval > self.base_interval:
                self._clean_streak = 0
                self.interval = max(self.base_interval, self.interval / 1.25)
                recovered = self.interval
        if recovered:
            print(f"  rate: 60 clean requests -- easing back up to "
                  f"~{60.0 / recovered:.1f} requests/min.")


# Process-wide singletons, configured once at the start of a run so every call
# site (list pages, detail panels, the recordsTotal probe) shares one budget.
LIMITER = None
THROTTLE = None


def configure_pacing(per_minute, floor_seconds, cooldown_seconds, trip_after=3):
    """Install the shared limiter + circuit breaker for this run."""
    global LIMITER, THROTTLE
    LIMITER = RateLimiter(per_minute=per_minute, floor_seconds=floor_seconds)
    THROTTLE = AdaptiveThrottle(trip_after=trip_after, cooldown_seconds=cooldown_seconds)
    print(f"  rate: global budget ~{per_minute:.0f} requests/min "
          f"({LIMITER.interval:.1f}s between requests, shared across ALL threads); "
          f"circuit breaker trips after {trip_after} failures -> {cooldown_seconds:.0f}s cooldown.")
    return LIMITER, THROTTLE


def request_with_retries(session, method, url, *, max_retries=8, timeout=30, **kwargs):
    """GET/POST that waits for a slot in the global rate budget, then retries with
    exponential backoff on network errors and 403/429/5xx.

    Honors Retry-After when the server sends it -- that is the server telling us
    exactly how long to wait, which beats guessing. Any pushback also feeds back
    into the shared limiter, so every other thread slows down too.
    """
    last_exc = None
    for attempt in range(1, max_retries + 1):
        if THROTTLE is not None:
            THROTTLE.wait_if_cooling_down()
        if LIMITER is not None:
            LIMITER.acquire()
        try:
            resp = session.request(method, url, timeout=timeout, **kwargs)
            if resp.status_code in (403, 429, 500, 502, 503, 504):
                raise requests.HTTPError(f"HTTP {resp.status_code}", response=resp)
            resp.raise_for_status()
            if LIMITER is not None:
                LIMITER.report_success()
            if THROTTLE is not None:
                THROTTLE.report_success()
            return resp
        except (requests.RequestException,) as exc:
            last_exc = exc
            resp_obj = getattr(exc, "response", None)
            status = getattr(resp_obj, "status_code", None)
            pushback = status in (403, 429, 503)

            retry_after = None
            if resp_obj is not None:
                ra_header = resp_obj.headers.get("Retry-After")
                if ra_header:
                    try:
                        retry_after = min(float(ra_header), 900)
                    except ValueError:
                        pass  # some servers send an HTTP-date instead of seconds -- skip those

            if pushback and LIMITER is not None:
                LIMITER.report_throttled()
            if THROTTLE is not None:
                THROTTLE.report_failure(hard=pushback, session=session)

            if attempt == max_retries:
                break

            # Recovering from a block needs far more patience than normal spacing:
            # start at 15s and climb toward 10 minutes.
            sleep_for = min(600.0, 15.0 * (2 ** (attempt - 1))) + random.uniform(0, 5)
            if retry_after is not None:
                sleep_for = max(sleep_for, retry_after)
            if LIMITER is not None:
                LIMITER.pause(sleep_for)  # everyone waits, not just this thread
            suffix = f" (server said Retry-After: {retry_after:.0f}s)" if retry_after else ""
            print(f"    ! {exc} -- retry {attempt}/{max_retries - 1} in {sleep_for:.1f}s{suffix}")
            time.sleep(sleep_for)
    raise RuntimeError(f"Request failed after {max_retries} attempts: {url}\n  {last_exc}")


class AdaptiveThrottle:
    """Thread-safe circuit breaker: if requests start failing across ANY of the
    worker threads (a strong signal the site has started rejecting us), pause ALL
    threads for a real cooldown instead of continuing to push a server that is
    already saying no. One success anywhere resets the counter.

    Two changes from the previous version:
      * it trips much sooner (3 failures, not 8) and cools down for minutes,
        not 45 seconds -- a WAF block is usually time-based, so the only thing
        that clears it is actually stopping;
      * each successive trip in the same run cools down longer than the last, and
        re-loads the search page afterwards so the session resumes with fresh
        cookies rather than replaying the exact request that just got refused.
    """

    def __init__(self, trip_after: int = 3, cooldown_seconds: float = 300.0):
        self._lock = threading.Lock()
        self._consecutive_failures = 0
        self.trip_after = trip_after
        self.cooldown_seconds = cooldown_seconds
        self.max_cooldown = max(cooldown_seconds * 6, 1800.0)
        self._cooldown_until = 0.0
        self._trips = 0

    def wait_if_cooling_down(self):
        while True:
            with self._lock:
                remaining = self._cooldown_until - time.monotonic()
            if remaining <= 0:
                return
            time.sleep(min(remaining, 5.0))  # wake periodically so Ctrl+C stays responsive

    def report_success(self):
        with self._lock:
            self._consecutive_failures = 0

    def report_failure(self, hard: bool = False, session=None):
        """hard=True means an explicit 403/429/503 (the server said no in words),
        which counts double against the trip threshold versus a plain timeout."""
        tripped = False
        cooldown = 0.0
        with self._lock:
            self._consecutive_failures += 2 if hard else 1
            if self._consecutive_failures >= self.trip_after:
                self._trips += 1
                cooldown = min(self.cooldown_seconds * (1.75 ** (self._trips - 1)),
                               self.max_cooldown)
                self._cooldown_until = time.monotonic() + cooldown
                self._consecutive_failures = 0
                tripped = True
        if tripped:
            resume = datetime.now().strftime('%H:%M:%S')
            print(f"  !! Server is refusing requests (trip #{self._trips}). Pausing ALL threads "
                  f"for {cooldown / 60:.1f} min from {resume} before trying again. "
                  f"Everything fetched so far is already cached -- Ctrl+C is safe.")
            if LIMITER is not None:
                LIMITER.pause(cooldown)
            if session is not None:
                time.sleep(min(cooldown, 5.0))
                warm_up_session(session, quiet=True)


def build_search_form(type_code, start, length, search_value="", draw=1):
    """Reproduce the DataTables POST body used by the site."""
    return {
        "draw": str(draw),
        "columns[0][data]": "CharityIPCName",
        "columns[0][name]": "",
        "columns[0][searchable]": "true",
        "columns[0][orderable]": "true",
        "columns[0][search][value]": "",
        "columns[0][search][regex]": "false",
        "columns[1][data]": "SectorAdministrato",
        "columns[1][name]": "",
        "columns[1][searchable]": "true",
        "columns[1][orderable]": "false",
        "columns[1][search][value]": "",
        "columns[1][search][regex]": "false",
        "columns[2][data]": "Activities",
        "columns[2][name]": "",
        "columns[2][searchable]": "true",
        "columns[2][orderable]": "false",
        "columns[2][search][value]": "",
        "columns[2][search][regex]": "false",
        "columns[3][data]": "RegistrationDate",
        "columns[3][name]": "",
        "columns[3][searchable]": "true",
        "columns[3][orderable]": "true",
        "columns[3][search][value]": "",
        "columns[3][search][regex]": "false",
        "order[0][column]": "0",
        "order[0][dir]": "asc",
        "start": str(start),
        "length": str(length),
        "search[value]": search_value,
        "search[regex]": "false",
        # This is the status filter. Passed as a literal string; requests will
        # url-encode it in the body exactly like the browser did.
        "query": f"?advType=0&type={type_code}",
        "sortColumn": "CharityIPCName",
        "sortDirection": "true",
        "reqType": "charityInfo",
        "filterColumn": "",
    }


def guid_to_query(guid: str) -> str:
    """The detail endpoint's `query` param is base64(GUID)."""
    return base64.b64encode(guid.encode()).decode()


def detail_url(guid: str, panel_label: str) -> str:
    """Build the detail URL, encoding the space in the panel name as %20 (not +)."""
    q = guid_to_query(guid)
    return f"{DETAIL_URL}?query={quote(q, safe='')}&type={quote(panel_label, safe='')}"


# --------------------------------------------------------------------------- #
# Cache helpers
# --------------------------------------------------------------------------- #

def cache_path(*parts) -> str:
    return os.path.join(CACHE_DIR, *parts)


def ensure_dirs():
    os.makedirs(CACHE_DIR, exist_ok=True)


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path, obj):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


# --------------------------------------------------------------------------- #
# Sub-command: selftest (offline)
# --------------------------------------------------------------------------- #

def cmd_selftest(_args):
    """Validate base64 / URL construction against the known CARITAS example -- no network."""
    guid = "d5916f1a-7382-e711-901b-005056962860"
    expected_q = "ZDU5MTZmMWEtNzM4Mi1lNzExLTkwMWItMDA1MDU2OTYyODYw"
    got_q = guid_to_query(guid)
    ok = got_q == expected_q
    print(f"GUID              : {guid}")
    print(f"base64(GUID)      : {got_q}")
    print(f"expected          : {expected_q}")
    print(f"base64 match      : {'PASS' if ok else 'FAIL'}")
    print()
    print("Profile URL       :", detail_url(guid, DETAIL_PANELS["profile"]))
    print("Financial URL     :", detail_url(guid, DETAIL_PANELS["financial"]))
    print()
    print("Search form sample (registered, first page):")
    form = build_search_form(STATUS_TYPE_CODES["registered"], start=0, length=5)
    print("  query =", form["query"])
    print()
    print("Selftest:", "PASS" if ok else "FAIL")
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# Sub-command: discover (which type code == which status)
# --------------------------------------------------------------------------- #

def cmd_discover(args):
    """Probe a range of `type` codes and report the CharityStatus each returns."""
    ensure_dirs()
    session = make_session()
    configure_pacing(per_minute=getattr(args, "rate", 12.0), floor_seconds=args.delay,
                     cooldown_seconds=180.0)
    warm_up_session(session)
    codes = list(range(args.start_code, args.start_code + args.count))
    print(f"Probing type codes {codes[0]}..{codes[-1]} (1 record each)\n")
    print(f"{'type code':>12} | {'recordsTotal':>12} | sample CharityStatus / Type")
    print("-" * 70)
    for code in codes:
        form = build_search_form(code, start=0, length=1)
        try:
            resp = request_with_retries(session, "POST", SEARCH_URL, data=form)
            data = resp.json()
            total = data.get("recordsTotal", "?")
            rows = data.get("charityInfosData") or []
            sample = ""
            if rows:
                r = rows[0]
                sample = f"{r.get('CharityStatus')} / {r.get('Type')}  e.g. {(''.join(str(r.get('CharityIPCName')))).strip()[:32]}"
            print(f"{code:>12} | {str(total):>12} | {sample}")
        except Exception as exc:
            print(f"{code:>12} | {'ERR':>12} | {exc}")
        time.sleep(args.delay)
    print("\nMap these to the statuses you need and update STATUS_TYPE_CODES at the top of the script.")
    return 0


# --------------------------------------------------------------------------- #
# Sub-command: scrape
# --------------------------------------------------------------------------- #

def dedupe_by_guid(rows):
    seen, out = set(), []
    for r in rows:
        gid = r.get("CharityAccountCRMRecordID")
        if gid and gid in seen:
            continue
        if gid:
            seen.add(gid)
        out.append(r)
    return out


def probe_total(session, type_code, timeout):
    """Cheap 1-row request just to learn the current recordsTotal."""
    form = build_search_form(type_code, start=0, length=1)
    resp = request_with_retries(session, "POST", SEARCH_URL, data=form, timeout=timeout)
    return resp.json().get("recordsTotal", 0)


def robust_probe_total(session, type_code, timeout, max_retries=6):
    """probe_total, but resilient: if the site is mid-throttle right when a
    resumed run starts, this retries with backoff instead of crashing before
    the main loop (which has its own, better-informed stall handling) even
    gets a chance to run."""
    for attempt in range(1, max_retries + 1):
        try:
            return probe_total(session, type_code, timeout)
        except Exception as exc:
            if attempt == max_retries:
                raise
            sleep_for = min(60, 3.0 * (2 ** (attempt - 1))) + random.uniform(0, 2)
            print(f"  list: couldn't check recordsTotal yet ({exc}) -- "
                  f"retry {attempt}/{max_retries - 1} in {sleep_for:.1f}s")
            time.sleep(sleep_for)


def fetch_list(session, status, type_code, page_size, delay, refresh=False,
                timeout=30, max_stall_retries=0):
    """Fetch the full charity list for a status, paginating. Cached AND resumable:
    if a previous run stopped early (rate-limited, interrupted, killed), this picks
    up from where it left off instead of starting over -- or, worse, silently
    trusting a cache that never got everything.

    A "stall" is a page that comes back empty/short even though more rows are
    still expected, OR a request that fails outright (including an explicit
    HTTP 429 -- confirmed to happen on this site after enough rapid requests).
    Both cases mean "the WAF is throttling us right now", not "the list is
    done" or "something is broken" -- so both retry with backoff through the
    SAME mechanism instead of one of them silently misreporting completion and
    the other crashing the whole script.
    """
    list_file = cache_path(f"list_{status}.json")

    all_rows = []
    if os.path.exists(list_file) and not refresh:
        all_rows = load_json(list_file)
        print(f"  list: found {len(all_rows)} cached rows -- verifying against live total...")

    total = robust_probe_total(session, type_code, timeout)
    print(f"  list: recordsTotal={total} (type={type_code}, page_size={page_size})")

    if all_rows and len(all_rows) >= total > 0:
        print(f"  list: cache already has all {len(all_rows)} rows -- nothing to fetch.")
        return all_rows

    if all_rows:
        print(f"  list: cache has {len(all_rows)}/{total} -- resuming from row {len(all_rows)}.")

    start = len(all_rows)
    draw = 0
    stall_retries = 0
    while start < total:
        draw += 1
        print(f"  list: requesting rows {start}-{min(start + page_size, total)} of {total} ...")
        form = build_search_form(type_code, start=start, length=page_size, draw=draw)

        request_failed = False
        fail_reason = None
        batch = []
        try:
            resp = request_with_retries(session, "POST", SEARCH_URL, data=form, timeout=timeout)
            data = resp.json()
            batch = data.get("charityInfosData") or []
        except Exception as exc:
            request_failed = True
            fail_reason = str(exc)

        reached_expected_end = (start + len(batch)) >= total
        is_stall = request_failed or not batch or (len(batch) < page_size and not reached_expected_end)

        if is_stall:
            # An explicit failure (e.g. HTTP 429) is a stronger throttling signal
            # than a merely-short page, so it escalates the backoff faster.
            stall_retries += 2 if request_failed else 1

            # A silently-empty page IS throttling on this site (confirmed: the WAF
            # returns 0 rows rather than an error once you have been too quick).
            # The old code only fed real exceptions back into the limiter, so the
            # exact failure mode you hit -- empty pages -- never slowed anything
            # down. Now it does, and it also counts toward tripping the breaker.
            if not request_failed:
                if LIMITER is not None:
                    LIMITER.report_throttled()
                if THROTTLE is not None:
                    THROTTLE.report_failure(hard=True, session=session)

            if max_stall_retries and stall_retries > max_stall_retries:
                print(f"  list: ** STALLED ** only got {len(all_rows)}/{total} rows after "
                      f"repeated throttling at start={start} ({fail_reason or 'empty pages'}). "
                      f"Saving what we have -- just re-run the same command later and it will "
                      f"resume automatically from here.")
                break

            # Backoff is deliberately decoupled from the request spacing: recovering
            # from a block needs minutes of genuine quiet, not seconds. Capped at
            # 15 min, and by default it never gives up (--max-stall-retries 0), since
            # progress is cached anyway -- an unattended run just outlasts the block.
            sleep_for = min(900.0, 20.0 * (2 ** min(stall_retries, 6))) + random.uniform(0, 10)
            reason = fail_reason if request_failed else (
                f"got {len(batch)} rows (expected up to {page_size}, {total - start} still remain)")
            budget = f"{stall_retries}/{max_stall_retries}" if max_stall_retries else f"attempt {stall_retries}"
            print(f"  list: {reason} -- throttled, waiting {sleep_for / 60:.1f} min "
                  f"before retrying ({budget}); progress so far is saved.")
            if LIMITER is not None:
                LIMITER.pause(sleep_for)
            time.sleep(sleep_for)

            # Re-establish the session before the next attempt: a fresh page load
            # is what a real user does after being told to go away.
            if stall_retries % 2 == 0:
                warm_up_session(session, timeout=timeout, quiet=True)
            continue

        stall_retries = 0
        all_rows.extend(batch)
        start += len(batch)
        save_json(list_file, dedupe_by_guid(all_rows))  # incremental -- survives Ctrl+C / crashes
        print(f"  list: {len(all_rows)}/{total} (saved)")
        if len(batch) < page_size:
            break  # genuinely the last page
        time.sleep(delay)

    deduped = dedupe_by_guid(all_rows)
    save_json(list_file, deduped)
    if total and len(deduped) < total:
        print(f"  list: ** INCOMPLETE ** saved {len(deduped)}/{total} rows -> {list_file}. "
              f"Re-run this same command to resume -- it will pick up where this left off.")
    else:
        print(f"  list: saved {len(deduped)} rows -> {list_file}")
    return deduped


def fetch_panel(session, guid, panel_key, delay, refresh=False, timeout=20, verbose=False, throttle=None):
    """Fetch one detail panel for one charity, cached per (guid, panel).
    Safe to call from multiple threads concurrently: cache reads/writes are
    per-(guid, panel) files, so different tasks never touch the same path."""
    path = cache_path("details", f"{guid}.{panel_key}.json")
    if os.path.exists(path) and not refresh:
        return load_json(path), True  # (payload, from_cache)

    if throttle is not None:
        throttle.wait_if_cooling_down()

    url = detail_url(guid, DETAIL_PANELS[panel_key])
    if verbose:
        print(f"    -> requesting {panel_key} panel: {url}")
    try:
        resp = request_with_retries(session, "GET", url, timeout=timeout)
    except Exception:
        if throttle is not None:
            throttle.report_failure()
        raise
    if throttle is not None:
        throttle.report_success()
    # The panel may return JSON or occasionally HTML/text; store whatever it is.
    try:
        payload = resp.json()
    except ValueError:
        payload = {"_raw_text": resp.text}
    save_json(path, payload)
    # No per-thread sleep here any more: pacing is owned by the global RateLimiter
    # inside request_with_retries, which is the only way to control the aggregate
    # rate correctly when several workers are running.
    return payload, False


def cmd_scrape(args):
    ensure_dirs()
    status = args.status
    if status not in STATUS_TYPE_CODES:
        print(f"Unknown status '{status}'. Choose from: {', '.join(STATUS_TYPE_CODES)}")
        return 2
    type_code = args.type_code if args.type_code is not None else STATUS_TYPE_CODES[status]
    timeout = getattr(args, "timeout", 30)
    verbose = getattr(args, "verbose", False)
    workers = max(1, getattr(args, "workers", 2))
    rate = max(1.0, getattr(args, "rate", 24.0))
    cooldown = max(30.0, getattr(args, "cooldown", 300.0))

    session = make_session(pool_size=max(workers, 8))
    print(f"[{datetime.now():%H:%M:%S}] Scraping status='{status}' "
          f"({STATUS_LABELS[status]}), type_code={type_code}")

    # One shared pacing budget for the entire process, then a real page load so
    # the session starts with cookies instead of cold-calling the XHR handler.
    configure_pacing(per_minute=rate, floor_seconds=args.delay, cooldown_seconds=cooldown)
    warm_up_session(session, timeout=timeout)

    est_minutes = (len(DETAIL_PANELS) * 2450) / rate
    print(f"  note: at ~{rate:.0f} req/min a full Registered run is roughly "
          f"{est_minutes / 60:.1f} hours of wall time. It is fully resumable and caches "
          f"as it goes, so leaving it running (or re-running it later) is the intended way "
          f"to finish -- going faster is what triggers the block.")

    # List-fetch stays single-threaded: it's inherently sequential pagination
    # (~50 requests total, not the bottleneck) and the stall-retry logic needs
    # a clear "one page in flight at a time" model.
    rows = fetch_list(session, status, type_code, args.page_size, args.delay,
                       refresh=args.refresh, timeout=timeout,
                       max_stall_retries=getattr(args, "max_stall_retries", 0))
    if args.limit:
        rows = rows[: args.limit]
        print(f"  (limited to first {len(rows)} charities via --limit)")

    # Build the work queue: one task per (charity, panel) still needing a fetch.
    # This is the expensive part (up to 2 x charities requests), so it's the
    # part that gets parallelized across --workers threads.
    tasks = []
    skipped = 0
    for rec in rows:
        guid = rec.get("CharityAccountCRMRecordID")
        name = ("".join(str(rec.get("CharityIPCName") or ""))).strip()
        if not guid:
            skipped += 1
            continue
        for panel_key in DETAIL_PANELS:
            tasks.append((guid, panel_key, name))

    total_charities = len(rows) - skipped
    print(f"[{datetime.now():%H:%M:%S}] Fetching Organisation Profile + Financial Information "
          f"for {total_charities} charities ({len(tasks)} requests total) "
          f"using {workers} parallel workers (skipped={skipped} with no GUID)...")

    throttle = THROTTLE
    counters = {"fetched": 0, "cached": 0, "errors": 0, "done": 0}
    lock = threading.Lock()
    n_tasks = len(tasks)

    def run_task(task):
        guid, panel_key, name = task
        try:
            _, from_cache = fetch_panel(session, guid, panel_key, args.delay, refresh=args.refresh,
                                         timeout=timeout, verbose=verbose, throttle=throttle)
            with lock:
                counters["cached" if from_cache else "fetched"] += 1
        except Exception as exc:
            with lock:
                counters["errors"] += 1
                print(f"  ERROR {panel_key} for {name[:40]}: {exc}")
        with lock:
            counters["done"] += 1
            done = counters["done"]
            if done % 40 == 0 or done == n_tasks:
                print(f"  [{datetime.now():%H:%M:%S}] [{done}/{n_tasks}] requests done "
                      f"(fetched={counters['fetched']}, cached={counters['cached']}, "
                      f"errors={counters['errors']})")

    if tasks:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [executor.submit(run_task, t) for t in tasks]
            for _ in as_completed(futures):
                pass  # run_task does all counting/printing/error-handling itself

    print(f"[{datetime.now():%H:%M:%S}] Scrape complete for '{status}': "
          f"{total_charities} charities, {counters['fetched']} panels fetched, "
          f"{counters['cached']} from cache, {counters['errors']} errors.")
    if args.export_after:
        return cmd_export(args)
    return 0


# --------------------------------------------------------------------------- #
# Sub-command: export (cache -> xlsx)
# --------------------------------------------------------------------------- #
#
# Column layout below is reverse-engineered from two REAL captured responses
# (not guessed): a "SearchOrgProfile" -> Organisation Profile payload and a
# Financial Information payload, both for CARITAS HUMANITARIAN AID & RELIEF
# INITIATIVES. If a future charity's response includes a field these dicts
# don't name, it is NOT dropped -- it lands in Profile_extra_<field> /
# Financial_extra_<field> on the Charities sheet so nothing is ever lost.

# Top-level keys in the Organisation Profile payload we handle explicitly.
PROFILE_HANDLED_KEYS = {
    "CharityName", "UENNo", "RegistrationDate", "ContactPerson", "OfficeNo", "FaxNo",
    "Email", "Website", "Address", "CharitySetup", "CharityCategory", "IPCStatus",
    "IPCPeriod", "shoIPCPeriod", "SingpassOFSCharityWithoutIPCOne", "SectorAdministrator",
    "LastProfileUpdate", "Objective", "VisionMission", "OrganisationActivities",
    "Patrons", "GoverningMembers", "KeyOfficers", "requiresLogin", "requiresDocLogin",
    "hide3Tabs", "FinancialSummary", "FinancialStatement", "ShowComplianceTab",
}

# Top-level keys in the Financial Information payload we handle explicitly.
FINANCIAL_HANDLED_KEYS = {
    "CharityName", "FinancialInfos", "requiresLogin", "requiresDocLogin",
    "FinancialInfoReceipts", "FinancialInfoExpenses", "FinancialInfoBalanceSheet",
    "FinancialInfoOtherInformation", "ShowComplianceTab", "ShowFSDetail",
}

# The four financial-breakdown tables share one pivoted shape: a JSON *string*
# holding a list of rows, where each row is a list of {"Key","Value"} pairs.
# The first pair is (table_label, line_item_name); every pair after that is
# (FY_period_label, amount) -- see FINANCIAL_TABLE_FIELDS below.
FINANCIAL_TABLE_FIELDS = [
    ("FinancialInfoReceipts", "Receipts"),
    ("FinancialInfoExpenses", "Expenses"),
    ("FinancialInfoBalanceSheet", "Balance Sheet"),
    ("FinancialInfoOtherInformation", "Other Information"),
]


def flatten(obj, prefix=""):
    """Flatten nested dict/list JSON into {dotted_key: scalar}. Used only as a
    catch-all for fields not covered by the explicit schema above, so unknown
    future fields still show up (as Profile_extra_* / Financial_extra_*)
    instead of silently vanishing."""
    out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            out.update(flatten(v, key))
    elif isinstance(obj, list):
        if all(not isinstance(x, (dict, list)) for x in obj):
            out[prefix] = "; ".join("" if x is None else str(x) for x in obj)
        else:
            out[prefix] = json.dumps(obj, ensure_ascii=False)
    else:
        out[prefix] = obj
    return out


def clean_cell(v):
    """Make a value safe/readable for an Excel cell.
    Lists of scalars -> '; '-joined; lists of objects / dicts -> JSON string."""
    if isinstance(v, list):
        if all(not isinstance(x, (dict, list)) for x in v):
            return "; ".join("" if x is None else str(x) for x in v)
        return json.dumps(v, ensure_ascii=False)
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    if isinstance(v, str):
        return v.replace("\u000b", " ").strip()
    return v


def strip_html(s):
    """Objective / VisionMission come back with <br/> tags and stray HTML."""
    if not isinstance(s, str):
        return s
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)
    s = re.sub(r"<[^>]+>", "", s)
    return s.strip()


def join_people(people):
    """[{'FullName':...,'Designation':...}, ...] -> one person per LINE:
        'NAME (Designation)'
        'NAME (Designation)'
    True multi-line text (paragraph style), not a semicolon-joined blob or JSON.
    Always returns a string ('' for empty/None), never None, so cells render blank
    not 'None'. Combine with wrap_text + a computed row height when writing the cell."""
    if not people:
        return ""
    lines = []
    for p in people:
        salu = (p.get("Salutation") or "").strip()
        name = (p.get("FullName") or "").strip()
        desig = (p.get("Designation") or "").strip()
        full_name = f"{salu} {name}".strip() if salu else name
        lines.append(f"{full_name} ({desig})" if desig else full_name)
    return "\n".join(lines)


def financial_summary_text(finfos):
    """FinancialInfos list -> one financial year per LINE, human-readable:
        'Oct 2024 - Sep 2025: Income $3,579,785 | Spending $3,330,149 (On Time)'
    Instead of the raw JSON blob the site's API actually returns."""
    if not finfos:
        return ""
    lines = []
    for rec in finfos:
        period = rec.get("FYPeriod") or "?"
        income = numify(rec.get("Income"))
        spending = numify(rec.get("Spending"))
        status = rec.get("Status") or ""
        income_s = f"${income:,.0f}" if isinstance(income, float) else "N/A"
        spending_s = f"${spending:,.0f}" if isinstance(spending, float) else "N/A"
        status_s = f" ({status})" if status else ""
        lines.append(f"{period}: Income {income_s} | Spending {spending_s}{status_s}")
    return "\n".join(lines)


def numify(v):
    """'3579785.0000' -> 3579785.0, but leaves blanks/non-numeric text alone."""
    if v in (None, ""):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def parse_financial_pivot(raw_json_str, table_label):
    """Turn one of the four FinancialInfo* pivoted strings into tidy rows:
    [{'Table':..., 'LineItem':..., 'FYPeriod':..., 'Value':...}, ...]"""
    if not raw_json_str:
        return []
    try:
        table_rows = json.loads(raw_json_str)
    except (TypeError, ValueError):
        return []
    out = []
    for row in table_rows:
        if not row:
            continue
        line_item = row[0].get("Value")
        for pair in row[1:]:
            out.append({
                "Table": table_label,
                "LineItem": line_item,
                "FYPeriod": pair.get("Key"),
                "Value": numify(pair.get("Value")),
            })
    return out


def cmd_export(args):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    status = args.status
    list_file = cache_path(f"list_{status}.json")
    if not os.path.exists(list_file):
        print(f"No cached list for '{status}'. Run `scrape` first ({list_file} missing).")
        return 2
    rows = load_json(list_file)
    if args.limit:
        rows = rows[: args.limit]

    charity_records = []   # one row per charity -> "Charities" sheet
    financial_rows = []    # one row per line item per FY -> "Financial Detail" sheet
    governance_rows = []   # one row per person -> "Governance" sheet

    for rec in rows:
        guid = rec.get("CharityAccountCRMRecordID")
        uen = rec.get("UENNo")
        name = ("".join(str(rec.get("CharityIPCName") or ""))).strip()

        flat = {}
        for k, v in rec.items():
            if k == "TotalRecords":
                continue  # same value on every row (just the pagination total) -- pure noise
            flat[k] = clean_cell(v)

        # ---- Organisation Profile panel -------------------------------------------------
        profile_path = cache_path("details", f"{guid}.profile.json")
        if os.path.exists(profile_path):
            p = load_json(profile_path)
            flat["Profile_ContactPerson"] = clean_cell(p.get("ContactPerson"))
            flat["Profile_OfficeNo"] = clean_cell(p.get("OfficeNo"))
            flat["Profile_FaxNo"] = clean_cell(p.get("FaxNo"))
            flat["Profile_Email"] = clean_cell(p.get("Email"))
            flat["Profile_Website"] = clean_cell(p.get("Website"))
            flat["Profile_Address"] = clean_cell(p.get("Address"))
            flat["Profile_CharitySetup"] = clean_cell(p.get("CharitySetup"))
            flat["Profile_CharityCategory"] = clean_cell(p.get("CharityCategory"))
            flat["Profile_IPCStatus"] = clean_cell(p.get("IPCStatus"))
            flat["Profile_IPCPeriod"] = clean_cell(p.get("IPCPeriod"))
            flat["Profile_SectorAdministrator"] = clean_cell(p.get("SectorAdministrator"))
            flat["Profile_LastProfileUpdate"] = clean_cell(p.get("LastProfileUpdate"))
            flat["Profile_Objective"] = strip_html(p.get("Objective"))
            flat["Profile_VisionMission"] = strip_html(p.get("VisionMission"))
            flat["Profile_OrganisationActivities"] = clean_cell(p.get("OrganisationActivities"))
            flat["Profile_GoverningBoard_Count"] = len(p.get("GoverningMembers") or [])
            flat["Profile_GoverningBoard_Names"] = join_people(p.get("GoverningMembers"))
            flat["Profile_KeyOfficers_Count"] = len(p.get("KeyOfficers") or [])
            flat["Profile_KeyOfficers_Names"] = join_people(p.get("KeyOfficers"))
            flat["Profile_Patrons_Count"] = len(p.get("Patrons") or [])
            flat["Profile_Patrons_Names"] = join_people(p.get("Patrons"))
            flat["Profile_RequiresLogin"] = clean_cell(p.get("requiresLogin"))
            flat["Profile_RequiresDocLogin"] = clean_cell(p.get("requiresDocLogin"))
            # Anything the schema above doesn't name yet -> never silently dropped.
            for k, v in p.items():
                if k in PROFILE_HANDLED_KEYS:
                    continue
                for fk, fv in flatten(v, f"Profile_extra_{k}").items():
                    flat[fk] = clean_cell(fv)

            # Governance sheet: one row per person.
            for role, people in (("Board Member", p.get("GoverningMembers")),
                                 ("Key Officer", p.get("KeyOfficers")),
                                 ("Patron", p.get("Patrons"))):
                for person in (people or []):
                    governance_rows.append({
                        "CharityIPCName": name, "UENNo": uen,
                        "Role": role,
                        "FullName": person.get("FullName"),
                        "Designation": person.get("Designation"),
                        "Salutation": person.get("Salutation"),
                    })
        else:
            flat["Profile_fetched"] = False

        # ---- Financial Information panel -------------------------------------------------
        financial_path = cache_path("details", f"{guid}.financial.json")
        if os.path.exists(financial_path):
            fin = load_json(financial_path)
            finfos = fin.get("FinancialInfos") or []
            for i in range(3):  # site itself only ever shows the latest 3 FYs
                rec_i = finfos[i] if i < len(finfos) else {}
                flat[f"Financial_FY{i+1}_Period"] = clean_cell(rec_i.get("FYPeriod"))
                flat[f"Financial_FY{i+1}_Income"] = numify(rec_i.get("Income"))
                flat[f"Financial_FY{i+1}_Spending"] = numify(rec_i.get("Spending"))
                flat[f"Financial_FY{i+1}_Status"] = clean_cell(rec_i.get("Status"))
            flat["Financial_Summary_Text"] = financial_summary_text(finfos)
            flat["Financial_ShowFSDetail"] = clean_cell(fin.get("ShowFSDetail"))
            flat["Financial_RequiresDocLogin"] = clean_cell(fin.get("requiresDocLogin"))
            for k, v in fin.items():
                if k in FINANCIAL_HANDLED_KEYS:
                    continue
                for fk, fv in flatten(v, f"Financial_extra_{k}").items():
                    flat[fk] = clean_cell(fv)

            # Financial Detail sheet: tidy long format, any number of years/line items.
            for field_name, table_label in FINANCIAL_TABLE_FIELDS:
                for item in parse_financial_pivot(fin.get(field_name), table_label):
                    financial_rows.append({
                        "CharityIPCName": name, "UENNo": uen, **item,
                    })
        else:
            flat["Financial_fetched"] = False

        charity_records.append(flat)

    # ---- Sheet: Charities (one row per charity) -------------------------------------------
    known_first = [
        "CharityIPCName", "UENNo", "Type", "CharityStatus", "IPCStatus",
        "IPCValidFrom", "IPCValidTill", "PrimarySector", "PrimaryClassification",
        "SectorAdministrato", "Activities", "RegistrationDate",
        "DeRegistrationDate", "DeRegistrationText", "CharitySetup",
        "NameofURL", "OtherURL",  # base list fields -- kept visible near the front, not buried at
                                   # the far right. Usually null in the source data for most
                                   # charities; that's a real data characteristic, not a bug.
    ]
    profile_cols = [
        "Profile_Website", "Profile_ContactPerson", "Profile_OfficeNo", "Profile_FaxNo",
        "Profile_Email", "Profile_Address", "Profile_CharitySetup", "Profile_CharityCategory",
        "Profile_IPCStatus", "Profile_IPCPeriod", "Profile_SectorAdministrator",
        "Profile_LastProfileUpdate", "Profile_Objective", "Profile_VisionMission",
        "Profile_OrganisationActivities", "Profile_GoverningBoard_Count",
        "Profile_GoverningBoard_Names", "Profile_KeyOfficers_Count", "Profile_KeyOfficers_Names",
        "Profile_Patrons_Count", "Profile_Patrons_Names",
        "Profile_RequiresLogin", "Profile_RequiresDocLogin",
    ]
    financial_cols = [
        "Financial_Summary_Text",
        "Financial_FY1_Period", "Financial_FY1_Income", "Financial_FY1_Spending", "Financial_FY1_Status",
        "Financial_FY2_Period", "Financial_FY2_Income", "Financial_FY2_Spending", "Financial_FY2_Status",
        "Financial_FY3_Period", "Financial_FY3_Income", "Financial_FY3_Spending", "Financial_FY3_Status",
        "Financial_ShowFSDetail", "Financial_RequiresDocLogin",
    ]
    tail_cols = ["CharityAccountCRMRecordID", "RelatedAccountOrgID"]

    fixed_order = known_first + profile_cols + financial_cols + tail_cols
    all_keys = []
    for r in charity_records:
        for k in r:
            if k not in all_keys:
                all_keys.append(k)

    def sort_key(k):
        if k in fixed_order:
            return (0, fixed_order.index(k))
        if k.startswith("Profile_extra_"):
            return (1, k)
        if k.startswith("Financial_extra_"):
            return (2, k)
        return (3, k)

    charity_columns = sorted(all_keys, key=sort_key)

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    body_font = Font(name="Arial")

    def write_sheet(ws, columns, records, wide_cols=(), multiline_cols=()):
        for c, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for r, rec in enumerate(records, 2):
            max_lines = 1
            for c, col in enumerate(columns, 1):
                val = rec.get(col, "")
                if isinstance(val, str) and len(val) > 32000:
                    val = val[:32000] + " …[truncated]"
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = body_font
                if col in multiline_cols:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if isinstance(val, str) and val:
                        max_lines = max(max_lines, val.count("\n") + 1)
            if max_lines > 1:
                # ~14pt per wrapped line reads cleanly at the default Arial 11 body font.
                ws.row_dimensions[r].height = min(14 * max_lines, 400)
        ws.freeze_panes = "A2"
        if columns:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
        for c, col in enumerate(columns, 1):
            letter = get_column_letter(c)
            if col in wide_cols or col in multiline_cols:
                width = 45
            else:
                width = min(max(len(col) + 2, 12), 28)
            ws.column_dimensions[letter].width = width

    ws1 = wb.active
    ws1.title = "Charities"
    multiline_cols = {
        "Profile_Objective", "Profile_VisionMission", "Profile_GoverningBoard_Names",
        "Profile_KeyOfficers_Names", "Profile_Patrons_Names", "Financial_Summary_Text",
    }
    write_sheet(
        ws1, charity_columns, charity_records,
        wide_cols={"CharityIPCName", "SectorAdministrato", "Activities", "Profile_Address",
                   "Profile_OrganisationActivities"},
        multiline_cols=multiline_cols,
    )

    ws2 = wb.create_sheet("Financial Detail")
    fin_cols = ["CharityIPCName", "UENNo", "Table", "LineItem", "FYPeriod", "Value"]
    write_sheet(ws2, fin_cols, financial_rows, wide_cols={"CharityIPCName", "LineItem"})

    ws3 = wb.create_sheet("Governance")
    gov_cols = ["CharityIPCName", "UENNo", "Role", "FullName", "Designation", "Salutation"]
    write_sheet(ws3, gov_cols, governance_rows, wide_cols={"CharityIPCName", "FullName"})

    out_path = args.out or f"{STATUS_LABELS[status].replace(' ', '_')}_enriched.xlsx"
    wb.save(out_path)
    print(f"Wrote {len(charity_records)} charities, {len(financial_rows)} financial-detail rows, "
          f"{len(governance_rows)} governance rows -> {out_path}")
    print(f"  Sheets: Charities ({len(charity_columns)} cols), Financial Detail, Governance")
    return 0


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def build_parser():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="command", required=True)

    sp = sub.add_parser("selftest", help="Offline validation of URL/base64 logic (no network).")
    sp.set_defaults(func=cmd_selftest)

    dp = sub.add_parser("discover", help="Probe type codes -> which CharityStatus each returns.")
    dp.add_argument("--start-code", type=int, default=100000000)
    dp.add_argument("--count", type=int, default=8)
    dp.add_argument("--delay", type=float, default=0.0)
    dp.add_argument("--rate", type=float, default=12.0,
                    help="Global requests per minute while probing (default 12).")
    dp.set_defaults(func=cmd_discover)

    common = dict()
    scp = sub.add_parser("scrape", help="Fetch list + both detail panels for a status (resumable, multithreaded).")
    scp.add_argument("--status", required=True, choices=list(STATUS_TYPE_CODES))
    scp.add_argument("--type-code", type=int, default=None,
                     help="Override the type code for this status (from `discover`).")
    scp.add_argument("--page-size", type=int, default=50, help="List page size (default 50).")
    scp.add_argument("--rate", type=float, default=24.0,
                     help="GLOBAL requests per minute across ALL threads (default 24, i.e. one "
                          "request every ~2.5s). This is the knob that actually decides whether "
                          "the site blocks you. Lower it (12, or 8) if you still get throttled.")
    scp.add_argument("--delay", type=float, default=0.0,
                     help="Minimum seconds between requests -- a floor under --rate. Usually leave "
                          "at 0 and tune --rate instead.")
    scp.add_argument("--cooldown", type=float, default=300.0,
                     help="Seconds all threads pause when the circuit breaker trips (default 300). "
                          "Each further trip in the same run waits longer.")
    scp.add_argument("--max-stall-retries", type=int, default=0, dest="max_stall_retries",
                     help="How many times to retry a throttled list page before giving up. "
                          "0 (default) = never give up; just keep waiting it out.")
    scp.add_argument("--workers", type=int, default=2,
                     help="Parallel worker threads for detail-panel fetching (default 2). Note that "
                          "workers no longer change the request RATE (--rate does) -- they only "
                          "overlap network latency, so raising this buys very little.")
    scp.add_argument("--timeout", type=float, default=30, help="Per-request timeout in seconds (default 30).")
    scp.add_argument("--limit", type=int, default=0, help="Only process first N charities (testing).")
    scp.add_argument("--refresh", action="store_true", help="Ignore cache and re-fetch.")
    scp.add_argument("--verbose", action="store_true",
                     help="Print every detail-panel request as it's sent (useful when diagnosing a stall).")
    scp.add_argument("--export-after", action="store_true", help="Run export when done.")
    scp.add_argument("--out", default=None, help="Output xlsx path (used with --export-after).")
    scp.set_defaults(func=cmd_scrape)

    ep = sub.add_parser("export", help="Build the enriched .xlsx from cache (no network).")
    ep.add_argument("--status", required=True, choices=list(STATUS_TYPE_CODES))
    ep.add_argument("--limit", type=int, default=0)
    ep.add_argument("--out", default=None, help="Output xlsx path.")
    ep.set_defaults(func=cmd_export)

    ap = sub.add_parser("all", help="scrape + export for one status.")
    ap.add_argument("--status", required=True, choices=list(STATUS_TYPE_CODES))
    ap.add_argument("--type-code", type=int, default=None)
    ap.add_argument("--page-size", type=int, default=50)
    ap.add_argument("--rate", type=float, default=24.0,
                    help="GLOBAL requests per minute across ALL threads (default 24).")
    ap.add_argument("--delay", type=float, default=0.0,
                    help="Minimum seconds between requests; a floor under --rate.")
    ap.add_argument("--cooldown", type=float, default=300.0)
    ap.add_argument("--max-stall-retries", type=int, default=0, dest="max_stall_retries")
    ap.add_argument("--workers", type=int, default=2)
    ap.add_argument("--timeout", type=float, default=30)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("--out", default=None)
    ap.set_defaults(func=lambda a: (setattr(a, "export_after", True) or cmd_scrape(a)))

    # scrape needs an export_after attribute even when called directly.
    scp.set_defaults(export_after=False)
    return p


def main(argv=None):
    args = build_parser().parse_args(argv)
    if not hasattr(args, "export_after"):
        args.export_after = getattr(args, "export_after", False)
    try:
        return args.func(args)
    except KeyboardInterrupt:
        print("\nInterrupted. Progress is cached incrementally, so just re-run the exact "
              "same command to resume -- nothing already fetched is lost.")
        return 130
    except Exception as exc:
        print(f"\n** Unexpected error: {exc}")
        print("Progress up to this point is cached (list rows and detail panels save "
              "incrementally), so re-running the exact same command will resume rather "
              "than start over. If this keeps happening, try a larger --delay and/or "
              "fewer --workers.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
